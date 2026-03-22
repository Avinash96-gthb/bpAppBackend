# ==============================
# bp_model.py  (UPDATED FINAL)
# ==============================

import numpy as np
from openpyxl import load_workbook


# -------------------------------------------------
# Bandpass filter for rPPG
# -------------------------------------------------
def bandpass(sig, fps, low=0.7, high=4):
    sig = np.asarray(sig, dtype=float)
    n = len(sig)
    if n < 8 or fps <= 0:
        return sig

    sig = sig - np.mean(sig)
    freqs = np.fft.rfftfreq(n, d=1.0 / fps)
    spec = np.fft.rfft(sig)
    mask = (freqs >= low) & (freqs <= high)
    spec[~mask] = 0
    filtered = np.fft.irfft(spec, n=n)
    return filtered


def _find_peaks(sig, min_distance):
    sig = np.asarray(sig, dtype=float)
    n = len(sig)
    if n < 3:
        return np.array([], dtype=int)

    min_distance = max(1, int(min_distance))
    candidate = []
    for idx in range(1, n - 1):
        if sig[idx] > sig[idx - 1] and sig[idx] >= sig[idx + 1]:
            candidate.append(idx)

    if not candidate:
        return np.array([], dtype=int)

    selected = []
    for idx in candidate:
        if not selected:
            selected.append(idx)
            continue

        if idx - selected[-1] >= min_distance:
            selected.append(idx)
        elif sig[idx] > sig[selected[-1]]:
            selected[-1] = idx

    return np.array(selected, dtype=int)


# -------------------------------------------------
# Extract mean green signal from ROI
# -------------------------------------------------
def extract_signal(frames, roi):
    x, y, w, h = roi

    signal = []
    for f in frames:
        fh, fw = f.shape[:2]
        x0 = max(0, int(x))
        y0 = max(0, int(y))
        x1 = min(fw, int(x + w))
        y1 = min(fh, int(y + h))

        if x1 <= x0 or y1 <= y0:
            continue

        patch = f[y0:y1, x0:x1]
        if patch.size == 0:
            continue

        g = np.mean(patch[:, :, 1])   # green channel
        signal.append(g)

    if not signal:
        raise ValueError("ROI produced empty signal. Please reselect ROI.")

    return np.array(signal, dtype=float)


# -------------------------------------------------
# Heart Rate from peaks
# -------------------------------------------------
def compute_hr(sig, fps):
    peaks = _find_peaks(sig, min_distance=fps / 2)

    if len(peaks) < 2:
        return 70  # fallback

    ibi = np.diff(peaks) / fps
    hr = 60 / np.mean(ibi)

    return hr


# -------------------------------------------------
# PTT calculation (ms)
# -------------------------------------------------
def compute_ptt(sig1, sig2, fps):
    p1 = _find_peaks(sig1, min_distance=fps / 2)
    p2 = _find_peaks(sig2, min_distance=fps / 2)

    if len(p1) == 0 or len(p2) == 0:
        return 150   # fallback

    delays = []

    for peak in p1:
        diff = p2 - peak
        diff = diff[diff > 0]
        if len(diff):
            delays.append(diff[0])

    if len(delays) == 0:
        return 150

    ptt_seconds = np.mean(delays) / fps

    return ptt_seconds * 1000   # convert → milliseconds


# -------------------------------------------------
# Train regression model
# -------------------------------------------------
def _read_training_rows(path="training_data.xlsx"):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("training_data.xlsx is empty")

    headers = [str(h).strip().upper() if h is not None else "" for h in rows[0]]
    required = ["PTT", "HR", "SYS", "DIA"]

    idx = {}
    for col in required:
        if col not in headers:
            raise ValueError(f"Missing required column '{col}' in training_data.xlsx")
        idx[col] = headers.index(col)

    data = []
    for row in rows[1:]:
        try:
            ptt = float(row[idx["PTT"]])
            hr = float(row[idx["HR"]])
            sys = float(row[idx["SYS"]])
            dia = float(row[idx["DIA"]])
            data.append((ptt, hr, sys, dia))
        except (TypeError, ValueError, IndexError):
            continue

    if len(data) < 3:
        raise ValueError("Need at least 3 valid rows in training_data.xlsx")

    return np.array(data, dtype=float)


def _fit_linear_2feat(X, y):
    X1 = np.column_stack([X, np.ones(len(X))])
    coef, _, _, _ = np.linalg.lstsq(X1, y, rcond=None)
    return coef


def load_model():
    data = _read_training_rows("training_data.xlsx")
    X = data[:, 0:2]  # PTT, HR
    y_sys = data[:, 2]
    y_dia = data[:, 3]

    sys_coef = _fit_linear_2feat(X, y_sys)
    dia_coef = _fit_linear_2feat(X, y_dia)
    return sys_coef, dia_coef


def _predict_linear(coef, ptt, hr):
    return float(coef[0] * ptt + coef[1] * hr + coef[2])


# -------------------------------------------------
# Main prediction function
# -------------------------------------------------
def predict_bp_from_frames(frames, fps, cheek_roi, palm_roi):

    # signals
    cheek_sig = extract_signal(frames, cheek_roi)
    palm_sig  = extract_signal(frames, palm_roi)

    # filter
    cheek_sig = bandpass(cheek_sig, fps)
    palm_sig  = bandpass(palm_sig, fps)

    # features
    hr  = compute_hr(cheek_sig, fps)
    ptt = compute_ptt(cheek_sig, palm_sig, fps)

    # load regression
    m_sys, m_dia = load_model()

    X = np.array([[ptt, hr]])   # IMPORTANT order

    sys_bp = _predict_linear(m_sys, X[0, 0], X[0, 1])
    dia_bp = _predict_linear(m_dia, X[0, 0], X[0, 1])

    return round(sys_bp,1), round(dia_bp,1), round(hr,1)